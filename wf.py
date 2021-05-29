from openpyxl import load_workbook
from openpyxl.styles import colors, Font, Fill, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment

def main():
    filename1="fri.xlsx"

    f2=load_workbook(filename1)
    names=f2.get_sheet_names()
    ws=f2[names[0]]

    rows = ws.max_row
    cols = ws.max_column

    # 字体
    font1 = Font(name='微软雅黑', size=11, b=True)
    font2 = Font(name='微软雅黑', size=11)

    # 边框
    line_t = Side(style='thin', color='000000')  # 细边框
    line_m = Side(style='medium', color='000000')  # 粗边框
    border0 = Border(right=line_t)
    border1 = Border(top=line_m, bottom=line_t, left=line_t, right=line_t)
    # 与标题相邻的边设置与标题一样
    border2 = Border(top=line_t, bottom=line_t, left=line_t, right=line_t)

    # fill = PatternFill('solid', fgColor='CFCFCF')
    alignment = Alignment(horizontal='right', vertical='center')
    alignment1 = Alignment(horizontal='center', vertical='center')


    try: 
        f2.add_named_style(NamedStyle(name='sty4',  border=border0, alignment=alignment))
    except: 
        print("enfo: sty4 exists")

    try:
        f2.add_named_style(NamedStyle(name='sty5',  border=border2, alignment=alignment))
    except:
        print("enfo: sty5 exists")

    try:
        f2.add_named_style(NamedStyle(name='sty6',  border=border2, alignment=alignment1))
    except:
        print("enfo: sty6 exists")




    data1=["","","2021/5/29"]+[i for i in range(25)]
    ws.append(data1)

    ws.cell(rows+1,2).alignment=Alignment(wrapText=True)
    ws.cell(rows+1,3).style="sty4"
    ws.cell(rows+1,8).style="sty4"
    ws.cell(rows+1,13).style="sty4"
    ws.cell(rows+1,18).style="sty4"
    ws.cell(rows+1,23).style="sty4"
    ws.cell(rows+1,28).style="sty4"

    ws1=f2[names[1]]

    rows1 = ws1.max_row
    cols1 = ws1.max_column

    data2=["2021/5/29"]+[i for i in range(10)]
    ws1.append(data2)

    ws1.cell(rows1+1,1).style="sty4"
    ws1.cell(rows1+1,6).style="sty4"
    ws1.cell(rows1+1,11).style="sty4"

    ws2=f2[names[2]]

    rows2 = ws2.max_row
    cols2 = ws2.max_column

    data3=["2021/5/29"]+["no error" for i in range(5)]
    ws2.append(data3)

    ws2.cell(rows2+1,1).style="sty5"
    # for i in range(2,cols2+1):
    for i in range(2,6+1):
        ws2.cell(rows2+1,i).style="sty6"

    f2.save(filename1)



if __name__=="__main__":
    main()